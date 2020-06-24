import itertools
import os

from openpyxl import load_workbook, Workbook


def get_directory_file_list(target_directory, file_list=[]):
    for i in os.listdir(target_directory):
        if os.path.isdir(target_directory + i + "/"):
            file_list = get_directory_file_list(target_directory + i + "/", file_list)
        else:
            file_list.append(target_directory + i)

    return file_list


def combination(file_name):
    nlp_list = []
    try:
        excel_filename = file_name
        load_wb = load_workbook(excel_filename, data_only=True)
        for sheet in load_wb:
            # 엑셀에 2열부터시작하고, 최대 2행만 가져오도록함
            for row in sheet.iter_rows(min_row=2, max_col=1):
                nlp_list.append(row[0].value)
    except Exception as e:
        print(e)
    com_list = list(map(','.join, itertools.combinations(nlp_list, 2)))

    if not os.path.isdir("./combination20" + str(file_name.split("/")[-1][0:2])):
        os.mkdir("./combination20" + str(file_name.split("/")[-1][0:2]))

    file_path = "./combination20" + str(file_name.split("/")[-1][0:2]) + "/" + file_name.split("/")[-1]
    write_wb = Workbook()
    write_ws = write_wb.active
    for i in range(0, len(com_list)):
        write_ws.cell(i + 1, 1, str(com_list[i].split(",")[0]))
        write_ws.cell(i + 1, 2, str(com_list[i].split(",")[1]))
        # print(str(noun_list_sort[i][0]) + " " + str(noun_list_sort[i][1]) + " 추가중")

    write_wb.save(file_path)


merge1_file_directory = "./merge1/"
file_list = get_directory_file_list(merge1_file_directory)
for i in file_list:
    combination(i)
