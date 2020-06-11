# 평균 시간 계산용 함수 time_spend 전역 배열을 사용
import csv
import os
import time
from datetime import datetime

from openpyxl import load_workbook, Workbook


def time_cal(count):
    sum = 0
    for i in time_spend:
        sum += i

    print("평균 소요 시간 : " + str(round(sum / (count + 1), 2)) + "초")


# 파일 처리에 필요한 폴더 생성
def init():
    if not os.path.isdir("./merge2"):
        os.mkdir("./merge2")


# 디렉토리의 파일들을 저장하는 배열을 생성
def get_directory_file_list(target_directory, file_list=[]):
    for i in os.listdir(target_directory):
        if os.path.isdir(target_directory + i + "/"):
            file_list = get_directory_file_list(target_directory + i + "/", file_list)
        else:
            file_list.append(target_directory + i)

    return file_list


def merge_2(file_number):
    # start1 = time.time()  # 시작 시간 저장
    # print(file_number + " 파일 처리 시작")
    # now = datetime.now()
    # process_starttime = now.strftime('%Y-%m-%d %H:%M:%S')
    # print("프로세스 처리 시작시간 " + process_starttime)  # 2020-05-29 22:49:32

    file_location = merge1_file_directory + file_number
    try:
        excel_filename = file_location
        load_wb = load_workbook(excel_filename, data_only=True)
        for sheet in load_wb:
            # 엑셀에 2열부터시작하고, 최대 2행만 가져오도록함
            for row in sheet.iter_rows(min_row=2, max_col=2):
                wtiv_no.append(file_number.split(".")[0])
                noun.append(row[0].value)
                nouncount.append(row[1].value)

    except Exception as e:
        print(e)

    # now = datetime.now()
    # process_endtime = now.strftime('%Y-%m-%d %H:%M:%S')
    # print("프로세스 처리 종료시간 " + process_endtime)  # 2020-05-29 22:49:32
    # time_spend.append(time.time() - start1)
    # print("파일처리 소요시간 : ", time.time() - start1)  # 현재시각 - 시작시간 = 실행 시간


def merge2_save(file_number):
    # start2 = time.time()  # 시작 시간 저장
    # now = datetime.now()
    # process_starttime = now.strftime('%Y-%m-%d %H:%M:%S')
    # print("파일 저장 시작 " + process_starttime)  # 2020-05-29 22:49:32

    # desc(내림차순으로 정렬)
    noun_list = {}
    for i in noun:
        noun_list.setdefault(i, nouncount[noun.index(i)])
    noun_list_sort = sorted(noun_list.items(), key=lambda x: x[1], reverse=True)

    file_path = merge1_file_directory + file_number + ".xlsx"
    write_wb = Workbook()
    write_ws = write_wb.active

    write_ws.cell(1, 1, "word")
    write_ws.cell(1, 2, "frequency")
    for i in range(0, len(noun_list_sort)):
        write_ws.cell(i + 2, 1, noun_list_sort[i][0])
        write_ws.cell(i + 2, 2, noun_list_sort[i][1])
        # print(str(noun_list_sort[i][0]) + " " + str(noun_list_sort[i][1]) + " 추가중")

    # write_wb.save(file_path)
    # now = datetime.now()
    # process_endtime = now.strftime('%Y-%m-%d %H:%M:%S')
    # print("파일 저장 완료 " + process_endtime)  # 2020-05-29 22:49:32
    # print("파일 저장 소요시간 : ", time.time() - start2)  # 현재시각 - 시작시간 = 실행 시간


start = time.time()  # 시작 시간 저장
wtiv_no = []
noun = []
nouncount = []
time_spend = []

init()
merge1_file_directory = "./merge1_2002/"
file_list = get_directory_file_list(merge1_file_directory)
for i in range(0, len(file_list)):
    merge_2(file_list[i].split("/")[-1])

noun_list = []
for i in range(0, len(noun)):
    templist = []
    templist.append(wtiv_no[i])
    templist.append(noun[i])
    templist.append(nouncount[i])
    noun_list.append(templist)

#print(noun_list)
csvf = open("./merge2/2002.csv", 'a', encoding='utf-8', newline='')
csvwrite = csv.writer(csvf)
csvwrite.writerow(["wtiv_no", "word", "frequency"])
csvf.close()
for i in noun_list:
    if start % 60 == 0:
        print("경과 시간 :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
    #print("총 " + str(len(noun_list)) + "행 중 " + str(noun_list.index(i)) + "번째 행 처리 중")
    #print(str(round((noun_list.index(i) / len(noun_list)) * 100, 2)) + "% 처리 중")
    #print("현재 파일 번호 : " + str(i[0]))
    csvf = open("./merge2/"+i[0][0:4] + ".csv", 'a', encoding='utf-8', newline='')
    csvwrite = csv.writer(csvf)
    csvwrite.writerow(i)
    csvf.close()

print("총 소요시간 :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
