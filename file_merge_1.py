from datetime import datetime
import os
import time

from openpyxl import load_workbook, Workbook


# 평균 시간 계산용 함수 time_spend 전역 배열을 사용
def time_cal(count):
    sum = 0
    for i in time_spend:
        sum += i

    print("평균 소요 시간 : " + str(round(sum / (count + 1), 2)) + "초")


# 파일 처리에 필요한 폴더 생성
def init():
    if not os.path.isdir("./merge1"):
        os.mkdir("./merge1")


# 디렉토리의 파일들을 저장하는 배열을 생성
def get_directory_file_list(target_directory, file_list=[]):
    for i in os.listdir(target_directory):
        if os.path.isdir(target_directory + i + "/"):
            file_list = get_directory_file_list(target_directory + i + "/", file_list)
        else:
            file_list.append(target_directory + i)

    return file_list


def merge_1(file_number):
    start1 = time.time()  # 시작 시간 저장
    print(file_number + " 파일 처리 시작")
    now = datetime.now()
    process_starttime = now.strftime('%Y-%m-%d %H:%M:%S')
    print("프로세스 처리 시작시간 " + process_starttime)  # 2020-05-29 22:49:32

    # complete폴더에있는 데이터를 활용하도록 함
    file_location = complete_file_directory + file_number
    try:
        excel_filename = file_location
        load_wb = load_workbook(excel_filename, data_only=True)
        for sheet in load_wb:
            # 엑셀에 2열부터시작하고, 최대 2행만 가져오도록함
            for row in sheet.iter_rows(min_row=2, max_col=2):
                if len(row[0].value) > 1:
                    #print(str(row[0].value) + " " + str(row[1].value) + " 추가중")
                    if row[0].value in noun:
                        nouncount[noun.index(row[0].value)] = int(nouncount[noun.index(row[0].value)]) + int(
                            row[1].value)
                    else:
                        noun.append(row[0].value)
                        nouncount.append(row[1].value)

    except Exception as e:
        print(e)

    try:
        except_word = open("결과제외단어.txt", mode='rt', encoding='utf-8')
        for i in except_word.readlines():
            i = i.replace('\n', '')
            del nouncount[noun.index(i)]
            del noun[noun.index(i)]
        except_word.close()
    except Exception as e:
        print(e)

    now = datetime.now()
    process_endtime = now.strftime('%Y-%m-%d %H:%M:%S')
    print("프로세스 처리 종료시간 " + process_endtime)  # 2020-05-29 22:49:32
    time_spend.append(time.time() - start1)
    print("파일처리 소요시간 : ", time.time() - start1)  # 현재시각 - 시작시간 = 실행 시간


def merge_save(file_number):
    start2 = time.time()  # 시작 시간 저장
    now = datetime.now()
    process_starttime = now.strftime('%Y-%m-%d %H:%M:%S')
    print("파일 저장 시작 " + process_starttime)  # 2020-05-29 22:49:32

    # desc(내림차순으로 정렬)
    noun_list = {}
    for i in noun:
        noun_list.setdefault(i, nouncount[noun.index(i)])
    noun_list_sort = sorted(noun_list.items(), key=lambda x: x[1], reverse=True)

    file_path = "./merge1/" + file_number + ".xlsx"
    write_wb = Workbook()
    write_ws = write_wb.active

    write_ws.cell(1, 1, "word")
    write_ws.cell(1, 2, "frequency")
    for i in range(0, len(noun_list_sort)):
        write_ws.cell(i + 2, 1, noun_list_sort[i][0])
        write_ws.cell(i + 2, 2, noun_list_sort[i][1])
        # print(str(noun_list_sort[i][0]) + " " + str(noun_list_sort[i][1]) + " 추가중")

    write_wb.save(file_path)
    now = datetime.now()
    process_endtime = now.strftime('%Y-%m-%d %H:%M:%S')
    print("파일 저장 완료 " + process_endtime)  # 2020-05-29 22:49:32
    print("파일 저장 소요시간 : ", time.time() - start2)  # 현재시각 - 시작시간 = 실행 시간


start = time.time()  # 시작 시간 저장
noun = []
nouncount = []
time_spend = []
previous_number = 0

init()
complete_file_directory = "./complete/"
file_list = get_directory_file_list(complete_file_directory)
# ex ) './complete/200101000012543_9.xlsx', './complete/200101001831944_2.xlsx'
# print(file_list)

try:
    for i in range(0, len(file_list) + 1):
        print("총 파일 " + str(len(file_list)) + "개 중 " + str(i) + "번째 파일 처리 중")
        print(str(round((i / len(file_list)) * 100, 2)) + "% 처리 중")
        # 기존의 파일 번호를 확인하고 틀리다면 저장하는 함수를 실행하고 파일처리를 하도록 함
        if int(file_list[i].split("/")[-1].split("_")[0]) != previous_number and previous_number != 0:

            print("현재번호 : " + str(file_list[i].split("/")[-1].split("_")[0]))
            print("기존번호 : " + str(previous_number))
            merge_save(file_list[i - 1].split("/")[-1].split("_")[0])
            noun = []
            nouncount = []
        previous_number = int(file_list[i].split("/")[-1].split("_")[0])
        merge_1(file_list[i].split("/")[-1])

        time_cal(i)
except:
    merge_save(file_list[-1].split("/")[-1].split("_")[0])

print("총 소요시간 :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
