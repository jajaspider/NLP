import os
import shutil
import time
import traceback
from io import StringIO

from bs4 import BeautifulSoup
from konlpy.tag import Okt
from openpyxl import Workbook
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

time_spend = []
okt = Okt()

def init():
    if not os.path.isdir("./complete"):
        os.mkdir("./complete")
    if not os.path.isdir("./temp"):
        os.mkdir("./temp")
    if not os.path.isdir("./text"):
        os.mkdir("./text")

# pdf 파일 주소를 넣으면 text로 반환해주는 함수
def read_pdf_PDFMINER(pdf_file_path):
    """
    pdf_file_path: 'dir/aaa.pdf'로 구성된 path로부터
    내부의 text 파일을 모두 읽어서 스트링을 리턴함.
    https://pdfminersix.readthedocs.io/en/latest/tutorials/composable.html
    """
    output_string = StringIO()
    with open(pdf_file_path, 'rb') as f:
        parser = PDFParser(f)
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        for page in PDFPage.create_pages(doc):
            interpreter.process_page(page)
    return str(output_string.getvalue())


# html 파일 주소를 넣으면 text로 반환해주는 함수
def html2text(html):
    soup = BeautifulSoup(open(html, encoding='UTF8'), 'html.parser')
    text_parts = soup.findAll(text=True)
    return ''.join(text_parts)


# 파일 확장자와 경로를 넣으면 자연어 처리 및 텍스트로 저장하는 함수
def txt_processing(extension, full_filename):
    # start1 = time.time()  # 시작 시간 저장
    file_name = full_filename.split("/")[-1]
    if os.path.isfile("./complete/" + file_name + ".xlsx"):
        return

    # 기록용 배열 2가지를 만듬
    noun = []
    nouncount = []
    just_text = ""

    # print("분석중인 파일 명 : " + file_name)
    # 한글 파일일 경우
    if extension == "hwp":
        try:
            # 한글파일을 html로 변환한다
            os.system("hwp5html.exe " + full_filename + " --output ./temp/" + file_name)

            htmlfile_path = "./temp/" + file_name + "/index.xhtml"
            # html에 있는 text를 받아온다
            just_text = html2text(htmlfile_path)
            # text를 임시로 text폴더에 저장한다
            txtfile_path = "./text/" + file_name + ".txt"
            f = open(txtfile_path, 'a', encoding='utf-8')
            f.write(just_text)
            f.close()
        except FileNotFoundError:
            print(file_name + " 파일없음")
            return
    if extension == "pdf":
        try:
            just_text = read_pdf_PDFMINER(full_filename)
            # text를 임시로 text폴더에 저장한다
            txtfile_path = "./text/" + file_name + ".txt"
            f = open(txtfile_path, 'a', encoding='utf-8')
            f.write(just_text)
            f.close()
        except:
            print("파일 읽기 불가")

    try:
        for j in okt.nouns(just_text):
            if j in noun:
                nouncount[noun.index(j)] += 1
            else:
                noun.append(j)
                nouncount.append(1)
    except:
        traceback.print_exc()

    noun_list = {}
    for j in noun:
        noun_list.setdefault(j, nouncount[noun.index(j)])
    noun_list_sort = sorted(noun_list.items(), key=lambda x: x[1], reverse=True)

    # d ntfs 년도 형식으로된것의 앞부분을 자르기위해 13번째부터 시작
    file_path = "./complete/" + file_name + ".xlsx"
    write_wb = Workbook()
    write_ws = write_wb.active

    write_ws.cell(1, 1, "word")
    write_ws.cell(1, 2, "frequency")
    for j in range(0, len(noun_list_sort)):
        write_ws.cell(j + 2, 1, noun_list_sort[j][0])
        write_ws.cell(j + 2, 2, noun_list_sort[j][1])

    write_wb.save(file_path)

    try:
        shutil.rmtree("./temp/" + file_name)
    except:
        traceback.print_exc()

    # print(file_name + " 처리완료")
    # time_spend.append(time.time() - start1)
    # print("소요시간 :", time.time() - start1)  # 현재시각 - 시작시간 = 실행 시간


def get_directory_file_list(target_directory, file_list=[]):
    for i in os.listdir(target_directory):
        if os.path.isdir(target_directory + i + "/"):
            file_list = get_directory_file_list(target_directory + i + "/", file_list)
        else:
            file_list.append(target_directory + i)

    return file_list

init()
file_list = get_directory_file_list("I:/nfds/2020/")
for i in range(0, len(file_list)):
    if os.path.isfile("./complete/" + file_list[i].split("/")[-1] + ".xlsx"):
        continue
    # print("총 파일 " + str(len(file_list)) + "개 중 " + str(i) + "번째 파일 처리 중")
    # print(str(round((i / len(file_list)) * 100, 2)) + "% 처리 중")
    if file_list[i].split("/")[-1] is not "2007413200031_2" or file_list[i].split("/")[-1] is not "200841180000004_4" or \
            file_list[i].split("/")[-1] is not "200845080000007_2" or file_list[i].split("/")[
        -1] is not "20074804000440_2" or file_list[i].split("/")[-1] is not "20074804000440_3" or \
            file_list[i].split("/")[-1] is not "20074804000440_4" or file_list[i].split("/")[
        -1] is not "20074804000440_5" or file_list[i].split("/")[-1] is not "20074804000440_6" or \
            file_list[i].split("/")[-1] is not "20074804000440_7" or file_list[i].split("/")[
        -1] is not "20074804000440_8" or file_list[i].split("/")[-1] is not "20074808000606_2":
        f = open(file_list[i], 'rb')
        metadata = f.read(16)
        #print(metadata)
        if metadata == b'\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01' or metadata == b'\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01' or metadata == b'\xff\xd8\xff\xe1t\x8cExif\x00\x00II*\x00' or metadata == b'\xff\xd8\xff\xe1\xff\xfeExif\x00\x00II*\x00' or metadata == b'\xff\xd8\xff\xe1QRExif\x00\x00II*\x00' or 'Exif' in str(
                metadata):
            print("JPEG")
        elif metadata[0:10] == b'%PDF-1.4\n%':
            # print("PDF")
            txt_processing("pdf", file_list[i])
        elif metadata == b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR':
            print("PNG")
        elif metadata[0:8] == b'PK\x03\x04\n\x00\x00\x00' or metadata[0:8] == b'PK\x03\x04\x14\x00\x00\x00':
            print("ZIP")
        elif metadata[0:8] == b'PK\x03\x04\x14\x00\x06\x00':
            print("XLSX")
        elif metadata == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1\x00\x00\x00\x00\x00\x00\x00\x00':
            # print("HWP")
            txt_processing("hwp", file_list[i])
        elif 'FasooSecure' in str(metadata):
            print("암호화 된 파일")
        elif 'EGG' in str(metadata):
            print("EGG")
        elif 'RIFF' in str(metadata):
            print("AVI")
