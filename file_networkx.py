# 디렉토리의 파일들을 저장하는 배열을 생성
import itertools
import operator
import os
import networkx as nx
from openpyxl import load_workbook
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import matplotlib


def get_directory_file_list(target_directory, file_list=[]):
    for i in os.listdir(target_directory):
        if os.path.isdir(target_directory + i + "/"):
            file_list = get_directory_file_list(target_directory + i + "/", file_list)
        else:
            file_list.append(target_directory + i)

    return file_list


def show_networkx(file_name):
    nlp_list = []
    G_centrality = nx.Graph()
    try:
        excel_filename = file_name
        load_wb = load_workbook(excel_filename, data_only=True)
        for sheet in load_wb:
            # 엑셀에 2열부터시작하고, 최대 2행만 가져오도록함
            for row in sheet.iter_rows(min_row=2, max_col=1):
                nlp_list.append(row[0].value)
    except Exception as e:
        print(e)
    com_list = list(map(','.join, itertools.combinations(nlp_list, 2)))
    print(com_list)
    for i in com_list:
        temp = i.split(",")
        G_centrality.add_edge(temp[0], temp[1])

    dgr = nx.degree_centrality(G_centrality)  # 연결 중심성
    btw = nx.betweenness_centrality(G_centrality)  # 매개 중심성
    cls = nx.closeness_centrality(G_centrality)  # 근접 중심성
    egv = nx.eigenvector_centrality(G_centrality)  # 고유벡터 중심성
    pgr = nx.pagerank(G_centrality)  # 페이지 랭크

    # 중심성이 큰 순서대로 정렬한다.
    sorted_dgr = sorted(dgr.items(), key=operator.itemgetter(1), reverse=True)
    sorted_btw = sorted(btw.items(), key=operator.itemgetter(1), reverse=True)
    sorted_cls = sorted(cls.items(), key=operator.itemgetter(1), reverse=True)
    sorted_egv = sorted(egv.items(), key=operator.itemgetter(1), reverse=True)
    sorted_pgr = sorted(pgr.items(), key=operator.itemgetter(1), reverse=True)
    print(sorted_dgr)
    print(sorted_btw)
    print(sorted_cls)
    print(sorted_egv)
    print(sorted_pgr)

    # 단어 네트워크를 그려줄 Graph 선언
    # G = nx.Graph()
    nlp_list = []

    # 페이지 랭크에 따라 두 노드 사이의 연관성을 결정한다. (단어쌍의 연관성)
    # 연결 중심성으로 계산한 척도에 따라 노드의 크기가 결정된다. (단어의 등장 빈도수)
    for i in range(len(sorted_pgr)):
        G.add_node(sorted_pgr[i][0], nodesize=sorted_dgr[i][1])

    try:
        excel_filename = file_name
        load_wb = load_workbook(excel_filename, data_only=True)
        for sheet in load_wb:
            # 엑셀에 2열부터시작하고, 최대 2행만 가져오도록함
            for row in sheet.iter_rows(min_row=2, max_col=1):
                nlp_list.append(row[0].value)
    except Exception as e:
        print(e)
    com_list = list(map(','.join, itertools.combinations(nlp_list, 2)))
    print(com_list)
    for i in com_list:
        temp = i.split(",")
        G.add_edge(temp[0], temp[1])




'''폰트설정
print ('버전: ', matplotlib.__version__)
print ('설치 위치: ', matplotlib.__file__)
print ('설정 위치: ', matplotlib.get_configdir())
print ('캐시 위치: ', matplotlib.get_cachedir())
print ('설정파일 위치: ', matplotlib.matplotlib_fname())
'''
G = nx.Graph()
merge1_file_directory = "./merge1/"
file_list = get_directory_file_list(merge1_file_directory)
for i in file_list:
    show_networkx(i)


options = {
        'edge_color': '#FFDEA2',
        'width': 1,
        'with_labels': True,
        'font_weight': 'regular',
}
# 노드 크기 조정
sizes = [G.nodes[node]['nodesize'] * 500 for node in G]
# 폰트 설정
fm._rebuild()  # 1회에 한해 실행해준다. (폰트 새로고침, 여러번 해줘도 관계는 없다.)
font_fname = './font/NanumSquareB.ttf'  # 여기서 폰트는 C:/Windows/Fonts를 참고해도 좋다.
fontprop = fm.FontProperties(fname=font_fname, size=18).get_name()
nx.draw(G, node_size=sizes, pos=nx.spring_layout(G, k=1, iterations=100), **options, font_family=fontprop)  # font_family로 폰트 등록
ax = plt.gca()
ax.collections[0].set_edgecolor("#555555")
plt.show()
